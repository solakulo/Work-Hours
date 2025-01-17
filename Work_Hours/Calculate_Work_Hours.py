import os
import json
from datetime import datetime, timedelta
import openpyxl

# 文件路径
INPUT_FILE = "system_logs_daily.txt"
CONFIG_FILE = "work_schedule_config.json"


def read_logs_from_txt(input_file):
    logs = []
    with open(input_file, "r", encoding="utf-8") as f:
        for line in f:
            parts = line.strip().split(", ")
            if len(parts) == 3:
                date, first_time, last_time = parts
                logs.append(
                    {"date": date, "first_time": first_time, "last_time": last_time}
                )
    return logs


def load_config():
    if os.path.exists(CONFIG_FILE):
        with open(CONFIG_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    else:
        raise FileNotFoundError("配置文件不存在，请先运行配置程序！")


def adjust_time_for_rest_periods(time, rest_periods, mode):
    """
    根据休息时间段调整时间。
    mode="start" 表示开机时间，mode="end" 表示关机时间。
    """
    for period in rest_periods:
        start, end = period.split("-")
        rest_start = datetime.strptime(start, "%H:%M")
        rest_end = datetime.strptime(end, "%H:%M")

        if mode == "start" and rest_start <= time < rest_end:
            return rest_end
        elif mode == "end" and rest_start < time <= rest_end:
            return rest_start

    return time


def calculate_work_and_overtime(log, config):
    work_start = datetime.strptime(log["first_time"], "%H:%M:%S")
    work_end = datetime.strptime(log["last_time"], "%H:%M:%S")

    # 调整时间根据休息时间段
    rest_periods = [config["lunch_break"]] + config["rest_periods"]
    work_start = adjust_time_for_rest_periods(work_start, rest_periods, mode="start")
    work_end = adjust_time_for_rest_periods(work_end, rest_periods, mode="end")

    # 计算工作时长
    work_duration = work_end - work_start

    # 减去午休和休息时间段
    for period in rest_periods:
        start, end = period.split("-")
        rest_start = datetime.strptime(start, "%H:%M")
        rest_end = datetime.strptime(end, "%H:%M")

        # 如果整个休息时间段在工作时间内，减去整个时间段
        if work_start <= rest_start and rest_end <= work_end:
            work_duration -= rest_end - rest_start
        # 如果部分休息时间段在工作时间内，只减去重叠部分
        elif work_start <= rest_start < work_end:
            overlap = min(work_end, rest_end) - rest_start
            if overlap > timedelta(0):
                work_duration -= overlap

    # 标准工作时长
    standard_start = datetime.strptime(config["standard_start_time"], "%H:%M")
    standard_end = datetime.strptime(config["standard_end_time"], "%H:%M")
    standard_duration = standard_end - standard_start

    # 计算加班时间
    overtime = max(timedelta(0), work_duration - standard_duration)

    return work_duration, overtime


def save_logs_to_excel_with_calculations(logs, config):
    yearly_logs = {}
    for log in logs:
        year = log["date"].split("-")[0]
        month = log["date"].split("-")[1]
        if year not in yearly_logs:
            yearly_logs[year] = {}
        if month not in yearly_logs[year]:
            yearly_logs[year][month] = []
        yearly_logs[year][month].append(log)

    for year, months in yearly_logs.items():
        output_excel = f"system_logs_{year}.xlsx"

        # 检查文件是否存在
        if os.path.exists(output_excel):
            wb = openpyxl.load_workbook(output_excel)
        else:
            wb = openpyxl.Workbook()
            default_sheet = wb.active
            wb.remove(default_sheet)  # 删除默认的 sheet1

        # 按月份创建或选择工作表
        for month, month_logs in months.items():
            if month not in wb.sheetnames:
                ws = wb.create_sheet(title=month)
                ws.append(
                    ["日期", "最早开机时间", "最晚关机时间", "工作时长", "加班时长"]
                )
            else:
                ws = wb[month]

            # 写入日志数据并计算
            for log in month_logs:
                if log["first_time"] == "无记录" or log["last_time"] == "无记录":
                    work_duration = "无记录"
                    overtime = "无记录"
                else:
                    work_duration, overtime = calculate_work_and_overtime(log, config)
                    work_duration = str(work_duration)
                    overtime = str(overtime)

                ws.append(
                    [
                        log["date"],
                        log["first_time"],
                        log["last_time"],
                        work_duration,
                        overtime,
                    ]
                )

        # 保存到文件
        wb.save(output_excel)
        print(f"日志已保存到 {output_excel}")


if __name__ == "__main__":
    try:
        config = load_config()
        logs = read_logs_from_txt(INPUT_FILE)
        if not logs:
            print("没有找到日志记录，请检查输入文件。")
        else:
            save_logs_to_excel_with_calculations(logs, config)
    except FileNotFoundError as e:
        print(e)
